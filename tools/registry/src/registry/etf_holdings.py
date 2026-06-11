"""Pull real ETF holdings (constituents + weights) from issuer files.

Confirmed-working primary sources (per research/results/Q8.md):
  - State Street SPDR — daily XLSX (11 Select Sector funds + SPY/DIA)
  - ARK — daily CSV (ARKK/ARKW/ARKG/ARKF/ARKQ/ARKX)
iShares/Invesco serve an HTML consent page to bare clients (left for later via a
session or aggregator fallback).

Holdings are normalized to {ticker, name, weight_pct, identifier} and disk-cached
per ETF (cache/etf/<TICKER>.json). We never redistribute the raw issuer file; we
compute derived weights by intersecting with our own registry (see match_holdings).
"""
import csv
import io
import json
import os
import re

CACHE_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "cache", "etf")
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124 Safari/537.36")

# rows that are cash / FX / derivatives, not tokenizable single stocks
_NON_EQUITY = {"", "-", "CASH", "USD", "MMF", "FUTURE", "SWAP"}
_TICKER_RE = re.compile(r"^[A-Z][A-Z.]{0,5}$")  # US single-stock ticker shape


def normalize_ticker(t):
    """Uppercase, trim, and fold class-share separators to '.' so the registry
    form (BRK-B) and issuer forms (BRK.B / BRK/B) all match."""
    if t is None:
        return ""
    return str(t).strip().upper().replace("/", ".").replace("-", ".")


def _pct(v):
    """Parse a weight cell: float, '10.15', or '10.15%'."""
    if v is None:
        return None
    s = str(v).strip().replace("%", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _keep(ticker):
    return ticker not in _NON_EQUITY and bool(_TICKER_RE.match(ticker))


def parse_ark_csv(text):
    """ARK daily CSV -> normalized holdings. Header:
    date,fund,company,ticker,cusip,shares,market value ($),weight (%)."""
    rows = list(csv.DictReader([l for l in text.splitlines() if l.strip()]))
    out = []
    for r in rows:
        ticker = normalize_ticker(r.get("ticker"))
        w = _pct(r.get("weight (%)"))
        if not _keep(ticker) or w is None:
            continue
        out.append({"ticker": ticker, "name": (r.get("company") or "").strip(),
                    "weight_pct": w, "identifier": (r.get("cusip") or "").strip()})
    return out


def parse_spdr_rows(rows):
    """SPDR holdings sheet (list of rows from openpyxl values_only) -> holdings.
    Finds the header row carrying 'Ticker' + 'Weight', then reads constituents."""
    hdr = None
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if "Ticker" in vals and any(v.startswith("Weight") for v in vals):
            hdr = i
            cols = vals
            break
    if hdr is None:
        return []
    ti = cols.index("Ticker")
    ni = cols.index("Name") if "Name" in cols else None
    wi = next(j for j, v in enumerate(cols) if v.startswith("Weight"))
    ii = cols.index("Identifier") if "Identifier" in cols else None
    out = []
    for row in rows[hdr + 1:]:
        if ti >= len(row):
            continue
        ticker = normalize_ticker(row[ti])
        w = _pct(row[wi]) if wi < len(row) else None
        if not _keep(ticker) or w is None:
            continue
        out.append({
            "ticker": ticker,
            "name": (str(row[ni]).strip() if ni is not None and ni < len(row) else ""),
            "weight_pct": w,
            "identifier": (str(row[ii]).strip() if ii is not None and ii < len(row) else ""),
        })
    return out


def match_holdings(holdings, registry_by_ticker):
    """Intersect holdings with the registry on ticker, renormalize weights across
    only the matched (tokenizable) subset. Returns (matched_constituents, coverage_pct)
    where coverage_pct = matched source-weight / total source-weight."""
    total = sum(h["weight_pct"] for h in holdings) or 1.0
    matched = [h for h in holdings if normalize_ticker(h["ticker"]) in registry_by_ticker]
    matched_w = sum(h["weight_pct"] for h in matched)
    coverage = round(100.0 * matched_w / total, 2)
    base = matched_w or 1.0
    out = []
    for h in matched:
        out.append({**h, "weight_pct": round(100.0 * h["weight_pct"] / base, 4)})
    return out, coverage


# ---- network adapters (not unit-tested; cached) ----

def _fetch(url):
    import requests
    r = requests.get(url, headers={"User-Agent": UA}, timeout=40)
    r.raise_for_status()
    return r


def fetch_spdr(ticker):
    url = ("https://www.ssga.com/library-content/products/fund-data/etfs/us/"
           f"holdings-daily-us-en-{ticker.lower()}.xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(_fetch(url).content), read_only=True, data_only=True)
    rows = [list(row) for row in wb.active.iter_rows(values_only=True)]
    return parse_spdr_rows(rows)


def fetch_ark(filename):
    # ARKQ/ARKX filenames contain a literal '&' that must be percent-encoded
    safe = filename.replace("&", "%26")
    url = f"https://assets.ark-funds.com/fund-documents/funds-etf-csv/{safe}"
    return parse_ark_csv(_fetch(url).text)


_ADAPTERS = {"spdr": fetch_spdr, "ark": fetch_ark}


def get_holdings(target, force=False):
    """Fetch (or load cached) normalized holdings for one target ETF config:
    {ticker, issuer, ark_file?}. Cached to cache/etf/<TICKER>.json."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    path = os.path.join(CACHE_DIR, f"{target['ticker']}.json")
    if not force and os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    issuer = target["issuer"]
    if issuer == "spdr":
        holdings = fetch_spdr(target["ticker"])
    elif issuer == "ark":
        holdings = fetch_ark(target["ark_file"])
    else:
        raise ValueError(f"unknown issuer {issuer}")
    with open(path, "w") as f:
        json.dump(holdings, f, ensure_ascii=False, indent=1)
    return holdings
